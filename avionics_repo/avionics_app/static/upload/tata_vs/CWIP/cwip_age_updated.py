from openpyxl import *
from datetime import datetime
import pandas as pd
import os
import fiscalyear

fiscalyear.START_MONTH=4

def find_age(matdocdate):
    date_format = "%d.%m.%Y"
    a = datetime.strptime('31.03.2022',date_format)
    b = datetime.strptime(matdocdate ,date_format)

    delta= a-b
    days=(delta.days)
    #print(days)
    yearrr = days/365
    #print(yearrr)
    year_fin=round(yearrr,2)
    return days,year_fin


path1=r"D:\Vendors\TATA_VS\CWIP_Report\data\input_data\lot3_dump1_wbs.xlsx"
wbl=load_workbook(path1)
wsl=wbl.active

path2=r"D:\Vendors\TATA_VS\CWIP_Report\data\input_data\merged_zps10_lot1.xlsx"
wb=load_workbook(path2)
ws=wb.active
zp_row=ws.max_row

zps10_unique_li=[]
for i in range(2,zp_row+1):
    zps=str(ws['D'+str(i)].value)
    if zps not in zps10_unique_li and zps != None:
        zps10_unique_li.append(zps)

zps10_unique_dict={}
for zp in zps10_unique_li:
    for i in range(2,zp_row+1):
        zps=str(ws['D'+str(i)].value)
        if zps==zp:
            budget_cost=ws['L'+str(i)].value
            stat_act_cost=ws['O'+str(i)].value
            zps10_unique_dict[zps]=[budget_cost,stat_act_cost]
print(zps10_unique_dict,len(zps10_unique_dict))
      
path4=r"D:\Vendors\TATA_VS\CWIP_Report\data\input_data\merged_excel_cn43n_lot3.xlsx"
wb3=load_workbook(path4)
ws3=wb3.active
x_row=ws3.max_row

wbs_unique=[]
for ii in range(1,x_row):
    wbs=str(ws3['D'+str(ii)].value)
    if wbs not in wbs_unique:
        wbs_unique.append(wbs)

print(len(wbs_unique),wbs_unique)
data_dict={}
for wbs1 in wbs_unique:
    for ii in range(1,x_row):
        wbs=str(ws3['D'+str(ii)].value)
        if wbs==wbs1:
            start_d=ws3['J'+str(ii)].value
            finish_d=ws3['K'+str(ii)].value
            data_dict[wbs1]=[start_d,finish_d]
print(len(data_dict),data_dict)
   
ws['A1']="Project def."
ws['B1']="Order"
ws['C1']="WBS element"
ws['D1']="RefDocNo"
ws['E1']="Cost Elem."
ws['F1']="Postg Date"
ws['G1']="Purch.Doc."
ws['H1']="Name"
ws['I1']="Material"
ws['J1']="Description"
ws['K1']="Quantity"
ws['L1']="PUM"
ws['M1']="Value TranCurr"
ws['N1']="User"
ws['O1']="Per"
ws['P1']="Year"
ws['Q1']="F.Y."
ws['R1']="Age in Days"
ws['S1']="Category"
ws['T1']="Upto 1-year"
ws['U1']="1-2 Year"
ws['V1']="2-3 Year"
ws['W1']="3+ year"

x=wsl.max_row
print(x)
cf=2022
cost_ele_li=[]
for j in range(2,x+1):
    cost_ele=str(wsl['E'+str(j)].value)
##    print(cost_ele)
    if cost_ele not in cost_ele_li and cost_ele!= "None" and cost_ele.startswith('4'):
        cost_ele_li.append(cost_ele)
        
print(len(cost_ele_li),cost_ele_li,"cost_ele_li")
output_excel_li=[]
for j in cost_ele_li:
    material_li=[]
    for i in range(2,x+1):
        cost_ele=str(wsl['E'+str(i)].value)
        if j==cost_ele:
            material=str(wsl['G'+str(i)].value)
            if material not in material_li and str(wsl['M'+str(i)].value)!="None":
                material_li.append(material)
    print(len(material_li),material_li,"material_li")
    for d in material_li:
##        print(d,j)
        quantity_sum=0
        value_sum=0
        k=1
        append_value_li=[]
        for i in range(2,x+1):
            cost_ele=str(wsl['E'+str(i)].value)
            material=str(wsl['G'+str(i)].value)
            post_date=wsl['M'+str(i)].value
            #print(post_date,i)
            if cost_ele==j and material==d and post_date!=None:
                #print(material,d,j)
                quantity=wsl['I'+str(i)].value
                value_trancurr=wsl['K'+str(i)].value
                project_def=wsl['C'+str(i)].value
                order_=wsl['D'+str(i)].value
                wbs_element=wsl['S'+str(i)].value
                ref_doc=wsl['L'+str(i)].value
                cost_ele=str(wsl['E'+str(i)].value)
                purch_doc=wsl['Q'+str(i)].value
                name=wsl['F'+str(i)].value
                material=wsl['G'+str(i)].value
                description=wsl['H'+str(i)].value
                quantity=wsl['I'+str(i)].value
                pum=wsl['J'+str(i)].value
                value_trancurr=wsl['K'+str(i)].value
                user=wsl['O'+str(i)].value
                per=wsl['P'+str(i)].value
                year=wsl['P'+str(i)].value
                xx=post_date.split('.')
                fy=fiscalyear.FiscalDate(int(xx[-1]),int(xx[1]),int(xx[0])).fiscal_year
##                print(fy)
                age=cf-fy
                fy=str(fy)
                days,age=find_age(post_date)
##                print(days,age)
                fiscal_year=str(year)+'-'+str(fy[2:])
                line_li=[project_def,order_,wbs_element,ref_doc,cost_ele,post_date,purch_doc,name,material,description,quantity,pum,value_trancurr,user,per,year,fiscal_year,days]
                if age<=1:
                    line_li.extend(["Upto 1-year",value_trancurr,'','',''])
                if age>1 and age<=2:
                    line_li.extend(["1-2 Year",'',value_trancurr,'',''])
                if age>2 and age<=3:
                    line_li.extend(["2-3 Year",'','',value_trancurr,''])
                if age>3:
                    line_li.extend(["3+ year",'','','',value_trancurr])
                    
                if quantity!= None :
                    quantity_sum=quantity_sum+quantity
                if  value_trancurr!= None:
                    value_sum=value_sum+value_trancurr
##                print(quantity,value_trancurr)
                k=k+1
                if order_ in data_dict:
                    start_li=[data_dict[order_][0],data_dict[order_][1]]
                    line_li.extend(start_li)
                else:
                    line_li.extend([None,None])
                append_value_li.append(line_li)
        print(k,j,quantity_sum,value_sum,"value_sum")
        if quantity_sum==0 and value_sum==0:
            print("value and quantity zero")
        else:
            df1=pd.DataFrame(append_value_li, columns = ["Project def.","Order","WBS element","RefDocNo","Cost Elem.","Postg Date","Purch.Doc.","Name","Material","Description","Quantity","PUM","Value TranCurr","User","Per","Year","F.Y.","Age in Days","Category","Upto 1-year","1-2 Year","2-3 Year","3+ year","start_date","finish_date"])
            output_excel_li.append(df1)
print(len(output_excel_li),output_excel_li)
if len(output_excel_li)>1:
    vertical_concat = pd.concat(output_excel_li, axis=0)
else:
    vertical_concat=output_excel_li[0]
path=r"G:\excel_basic"
vertical_concat.to_excel(os.path.join(path,'lot3_age.xlsx'),index=False)
